# coding: utf-8

import email, smtplib, ssl

from flask_restful import Resource, MethodView
from flask import json, request, jsonify, send_file
from db import deleteInStore, executeSql, insertIntoStore
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def getStoreInformation(store_id):
    store = executeSql("SELECT * FROM store WHERE id = {};".format(store_id))[0]
    newStore = dict()
    newStore['Nome'] = store.get('nome') + ' - ' + store.get('logradouro')
    newStore['Endereco'] = store.get('endereco')
    newStore['Telefone'] = store.get('telefone')
    newStore['Horarios'] = store.get('horarios')
    return newStore

def generateExcel(store, post_data):
    wb = Workbook()
    del wb['Sheet']
    # Create second sheet on workbook
    ws1 = wb.create_sheet("Produtos")

    ws1.append(['Corredor', 'Subcorredor', 'Nome do Produto', 'Descrição do Produto', 'Preço do Produto'])
    for value in post_data:
        if not value:
            continue
        ws1.append([value.get('corredor_template'), value.get('sub_corredor_template'), value.get('nome_do_produto_template'), value.get('descricao_do_produto_template'), value.get('preco_do_produto_template').replace(',', '.')])

    # Create second sheet on workbook
    ws2 = wb.create_sheet("Dados da loja")

    ws2.append(['Nome', 'Endereco', 'Telefone', 'Horarios'])
    ws2.append([store['Nome'], store['Endereco'], store['Telefone'], store['Horarios']])
    
    wb.save(store['Nome'].replace(' ', '').replace('-', '_') + '.xlsx')
    wb.close()


def send_mail(store):
    subject = "An email with attachment from Python"
    body = "This is an email with attachment sent from Python"
    sender_email = "napp.mail.to.rappi@gmail.com"
    receiver_email = "ygor@napp.ws" # Set who will receive the e-mail
    password = 'napp154878'

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Bcc"] = receiver_email  # Recommended for mass emails

    # Add body to email
    message.attach(MIMEText(body, "plain"))

    filename = store['Nome'].replace(' ', '').replace('-', '_') + '.xlsx'  # In same directory as script

    # Open XLSX file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        "attachment; filename= {}".format(filename),
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()

    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, text)

def post(store_id):
    # Store information
    store = getStoreInformation(store_id)

    # Get post data
    post_data = request.get_json()

    # Generate excel
    print('Generating excel...')
    generateExcel(store, post_data)
    
    # Send email
    print('Sending email...')
    send_mail(store)
        
        
